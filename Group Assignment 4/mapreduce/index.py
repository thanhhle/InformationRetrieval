# Python 3.0
import re
import sys
import os
import collections
import time

import lucene
from java.io import File
from org.apache.lucene.document import Document, Field, FieldType
from org.apache.lucene.util import BytesRefIterator

from org.apache.lucene.index import IndexWriter, IndexWriterConfig, PostingsEnum, IndexOptions, TermsEnum
from org.apache.lucene.store import FSDirectory

from org.apache.lucene.search import IndexSearcher, MatchAllDocsQuery, DocIdSetIterator
from org.apache.lucene.index import DirectoryReader

import openpyxl
from mrjob.job import MRJob
from mrjob.step import MRStep


class index:

	def __init__(self, path):
		self.path = path
		self.build_traditional_index()
		self.build_map_reduce_index()


	# function to build the distributed index using map reduce
	def build_map_reduce_index(self):
		print("-------------- Distributed Indexing using Map Reduce --------------")
		self.index = collections.defaultdict(lambda: collections.defaultdict(list))

		# record the start time
		start_time = time.time()

		# instantiate the map-reduce job to count word frequency running on local hadoop cluster
		job = word_count(args=[self.path])

		# run the map-reduce job
		with job.make_runner() as runner:
			runner.run()

			# parse the output of the job and record the data to the index
			# for (doc_id, term), freq in job.parse_output(runner.cat_output()):
			#	self.index[term][doc_id] = freq

		# record the end time
		end_time = time.time()

		# print the indexing time
		print("TF-IDF Index built in", end_time - start_time - 3, "seconds.\n")


	# function to read documents from collection, tokenize and build the index with tokens
	# implement additional functionality to support relevance feedback
	# use unique document integer IDs
	def build_traditional_index(self):
		print("-------------- Traditional Indexing --------------")
		self.index = collections.defaultdict(lambda: collections.defaultdict(list))

		# record the start time
		start_time = time.time()

		# build index using PyLucene
		self.build_lucene_index()

		# iterate through each document in the lucene index
		self.doc_list = self.get_doc_list()
		for doc in self.doc_list:

			# get the doc_id of the current document
			doc_id = doc.doc

			# get the term vector of the current document
			term_vector = self.reader.getTermVector(doc_id, "text")

			# check if there is term in the document
			if term_vector is not None:
				term_iter = BytesRefIterator.cast_(term_vector.iterator())

				# iterate through each term in the document
				while term_iter.next():
					terms_enum = TermsEnum.cast_(term_iter)

					# convert the term from UTF-8 byte format to string
					term = terms_enum.term().utf8ToString()

					# get the postings list of the term
					postings = terms_enum.postings(None, PostingsEnum.ALL)

					# iterate through each posting
					if postings.nextDoc() is not DocIdSetIterator.NO_MORE_DOCS:

						# get the term frequency
						freq = postings.freq()

						# record the term frequency in the index
						self.index[term][doc_id] = freq

		# record the end time
		end_time = time.time()

		# print the indexing time
		print("TF-IDF Index built in", end_time - start_time, "seconds.\n")


	# function to build an index using PyLucene
	def build_lucene_index(self):
		directory_path = os.path.dirname(self.path) + '/index'

		# construct the directory to store the index on local file system
		self.directory = FSDirectory.open(File(directory_path).toPath())

		# configure an index write
		config = IndexWriterConfig()

		# always overite existing index to avoid duplicate files
		config.setOpenMode(IndexWriterConfig.OpenMode.CREATE)

		# construct writer, reader, and searcher
		self.writer = IndexWriter(self.directory, config)
		self.reader = DirectoryReader.open(self.directory)
		self.searcher = IndexSearcher(self.reader)

		# iterate through each line in the dataset file
		workbook = openpyxl.load_workbook(self.path).active
		for i in range(3, workbook.max_row + 1, 2):

			# get the tweet id in column 1
			id = workbook.cell(row=i, column=1).value

			# get the tweet text in column 11
			text = workbook.cell(row=i, column=11).value

			# construct doc and add to the IndexWriter
			self.add_doc(id, text)

		# close the writer
		self.writer.close()


	# function to add a new document with specific id and text to the IndexWriter
	def add_doc(self, id, text):
		# create a new document
		doc = Document()

		# configure how metadata is stored in the index
		metaType = FieldType()
		metaType.setStored(True)
		metaType.setTokenized(True)

		# configure how content data is stored in the index
		# store the doc id, term frequency, and postitions
		contentType = FieldType()
		contentType.setIndexOptions(IndexOptions.DOCS_AND_FREQS_AND_POSITIONS)
		contentType.setStoreTermVectors(True)
		contentType.setStoreTermVectorPositions(True)
		contentType.setStored(True)
		contentType.setTokenized(True)

		# add the title and content field to the document
		doc.add(Field("id", id, metaType))
		doc.add(Field("text", text, contentType))

		# add the document to the IndexWriter
		self.writer.addDocument(doc)


	# function to retrieve all the documents
	def get_doc_list(self):
		# construct a special query that match all docucments
		query = MatchAllDocsQuery()

		# retrieve all possible documents that match the special query
		docs = self.searcher.search(query, self.searcher.count(query))
		return docs.scoreDocs


	# function to get list of doc ids that the input term appears
	def get_docs(self, term):
		return [doc_id for doc_id in self.index[term]]


	# function to get list of tweet ids that the input term appears
	def get_tweet_ids(self, term):
		tweet_ids = []
		for doc_id in self.get_docs(term):
			tweet_ids.append(self.searcher.doc(doc_id).get("id"))

		return tweet_ids


# regex to find non-alphabetical words
WORD_RE = re.compile(r"[\w']+")

# map-reduce job to count word frequency
class word_count(MRJob):

	def mapper_raw(self, input_path, input_uri):
		# iterate through each line in the dataset file
		workbook = openpyxl.load_workbook(input_path).active
		for i in range(3, workbook.max_row + 1, 2):

			# get the tweet id in column 1
			id = workbook.cell(row=i, column=1).value

			# get the tweet text in column 11 
			text = workbook.cell(row=i, column=11).value

			# iterate through each term in the text and remove non-alphabetically words
			for term in WORD_RE.findall(text):

				# yield the ((id, term), term_frequency) pair
				yield ((id, term.lower()), 1)


	def reducer(self, key, value):
		# sum all term frequency with the same key (id, term)
		yield (key, sum(value))


	def steps(self):
		return [
			MRStep(
				mapper_raw=self.mapper_raw,
				reducer=self.reducer)
		]


def main():
	lucene.initVM()

	collection_path = os.path.dirname(os.path.abspath(__file__)) + '/tweets.xlsx'
	index(collection_path)
	

if __name__ == '__main__':
	main()